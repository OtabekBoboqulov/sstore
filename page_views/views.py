import cloudinary.uploader
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import Sum, Case, When, IntegerField
from datetime import datetime
from django.utils import timezone
from django.http import HttpResponse
from api.serializers import ProductSerializer, CategorySerializer, ProductUpdateSerializer, ExpenseSerializer, \
    MarketSerializer, DebtorSerializer
from api.authentication import CustomTokenAuthentication
from products.models import Product, Category, ProductUpdate
from reports.models import Expense, Debtor
from markets.models import Market
from io import BytesIO
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import pandas as pd


def order_products_by_sells(markets):
    return Product.objects.filter(category_id__in=markets).annotate(
        total_subtracted=Sum(
            Case(
                When(updates__status='subed', then='updates__quantity'),
                default=0,
                output_field=IntegerField()
            )
        )
    ).order_by('-total_subtracted')


def order_products_by_price(markets):
    return Product.objects.filter(category_id__in=markets).annotate(
        total_sold_price=Sum(
            Case(
                When(updates__status='subed', then='updates__price'),
                default=0,
                output_field=IntegerField()
            )
        )
    ).order_by('-total_sold_price')


@api_view(['GET'])
@authentication_classes([CustomTokenAuthentication])
@permission_classes([IsAuthenticated])
def dashboard(request):
    categories = Category.objects.all().filter(market_id=request.user.id)
    category_serialized = CategorySerializer(categories, many=True)
    markets = [category['id'] for category in category_serialized.data if category['market_id'] == request.user.id]
    products = Product.objects.all().filter(category_id__in=markets)
    products_serialized = ProductSerializer(products, many=True)
    product_ids = [product['id'] for product in products_serialized.data]
    current_month = datetime.now().month
    current_month_text = datetime.now().strftime('%B')
    current_day = datetime.now().day
    start_date = timezone.make_aware(datetime(datetime.now().year, current_month, 1))
    profit = list()
    market_data = request.user
    market_serialized = MarketSerializer(market_data)
    for i in range(1, current_day + 1):
        end_date = timezone.make_aware(datetime(datetime.now().year, current_month, i, 23, 59, 59, 999999))
        updates = ProductUpdate.objects.all().filter(product_id__in=product_ids, status='subed', date__gte=start_date,
                                                     date__lte=end_date)
        updates_added = ProductUpdate.objects.all().filter(product_id__in=product_ids, status='added',
                                                           date__gte=start_date, date__lte=end_date)
        updates_serialized = ProductUpdateSerializer(updates, many=True)
        updates_added_serialized = ProductUpdateSerializer(updates_added, many=True)
        profit_from_sells = sum(list(map(float, [update['price'] for update in updates_serialized.data])))
        expanse_from_products = sum(list(map(float, [update['price'] for update in updates_added_serialized.data])))
        expanses = Expense.objects.all().filter(market_id=request.user.id, date__gte=start_date, date__lte=end_date)
        expanses_serialized = ExpenseSerializer(expanses, many=True)
        expanses_sum = sum(list(map(float, [expanse['price'] for expanse in expanses_serialized.data])))
        profit.append(profit_from_sells - expanse_from_products - expanses_sum)
    try:
        income = profit_from_sells
        expanses_total = expanses_sum + expanse_from_products
        products_added = sum(list(map(int, [update['quantity'] for update in updates_added_serialized.data])))
        products_subbed = sum(list(map(int, [update['quantity'] for update in updates_serialized.data])))
    except NameError:
        income, expanses_total, products_subbed, products_added = 0, 0, 0, 0
    products_ordered_by_sells = order_products_by_sells(markets)
    products_serialized_by_sells = ProductSerializer(products_ordered_by_sells, many=True)
    products_ordered_by_price = order_products_by_price(markets)
    products_serialized_by_price = ProductSerializer(products_ordered_by_price, many=True)
    quantity = sum([product["quantity"] for product in products_serialized.data])
    return Response([{'products': products_serialized.data}] + [{'quantity': quantity}] + [
        {'products_by_sells': products_serialized_by_sells.data}] + [
                        {'products_by_price': products_serialized_by_price.data}] + [{'profit': profit}] + [
                        {'market_data': market_serialized.data}] + [{'income': income}] + [
                        {'expanses_total': expanses_total}] + [{'products_subbed': products_subbed}] + [
                        {'products_added': products_added}] + [{'current_month': current_month_text}])


@api_view(['GET'])
@authentication_classes([CustomTokenAuthentication])
@permission_classes([IsAuthenticated])
def categories(request):
    categories = Category.objects.all().filter(market_id=request.user.id)
    category_serialized = CategorySerializer(categories, many=True)
    return Response(category_serialized.data)


@api_view(['POST'])
@authentication_classes([CustomTokenAuthentication])
@permission_classes([IsAuthenticated])
def category_create(request):
    try:
        category = Category(name=request.data['name'], market_id=request.user)
        category.save()
        return Response({'message': 'Category created successfully'})
    except Exception as e:
        return Response({'error': str(e)}, status=400)


@api_view(['DELETE'])
@authentication_classes([CustomTokenAuthentication])
@permission_classes([IsAuthenticated])
def category_delete(reqeust, pk):
    category = Category.objects.get(id=pk)
    category.delete()
    return Response({'message': 'Category deleted successfully'})


@api_view(['GET'])
@authentication_classes([CustomTokenAuthentication])
@permission_classes([IsAuthenticated])
def categories_with_products(request):
    categories = Category.objects.all().filter(market_id=request.user.id)
    result = {}
    for category in categories:
        result[str(category)] = ProductSerializer(category.products, many=True).data
    return Response(result)


@api_view(['GET'])
@authentication_classes([CustomTokenAuthentication])
@permission_classes([IsAuthenticated])
def products(request):
    categories = Category.objects.all().filter(market_id=request.user.id)
    category_serialized = CategorySerializer(categories, many=True)
    markets = [category['id'] for category in category_serialized.data if category['market_id'] == request.user.id]
    products = Product.objects.all().filter(category_id__in=markets)
    products_serialized = ProductSerializer(products, many=True)
    products_quantity = len(products_serialized.data)
    available_products = len([product for product in products_serialized.data if product['status'] == 'available'])
    few_products = len([product for product in products_serialized.data if product['status'] == 'few'])
    ended_products = len([product for product in products_serialized.data if product['status'] == 'ended'])
    return Response({'products': products_serialized.data, 'products_quantity': products_quantity,
                     'available_products': available_products, 'few_products': few_products,
                     'ended_products': ended_products})


@api_view(['GET'])
@authentication_classes([CustomTokenAuthentication])
@permission_classes([IsAuthenticated])
def product_detail(request, pk):
    product = Product.objects.get(id=pk)
    product_serialized = ProductSerializer(product)
    current_month = datetime.now().month
    start_date = timezone.make_aware(datetime(datetime.now().year, current_month, 1))
    product_sold = ProductUpdate.objects.all().filter(product_id=pk, date__gte=start_date, status='subed')
    product_sold_serialized = ProductUpdateSerializer(product_sold, many=True)
    total_sold = sum(list(map(float, [update['price'] for update in product_sold_serialized.data])))
    product_bought = ProductUpdate.objects.all().filter(product_id=pk, date__gte=start_date, status='added')
    product_bought_serialized = ProductUpdateSerializer(product_bought, many=True)
    total_bought = sum(list(map(float, [update['price'] for update in product_bought_serialized.data])))
    return Response({'product': product_serialized.data, 'product_sold': product_sold_serialized.data,
                     'product_bought': product_bought_serialized.data, 'total_sold': total_sold,
                     'total_bought': total_bought})


@api_view(['GET'])
@authentication_classes([CustomTokenAuthentication])
@permission_classes([IsAuthenticated])
def product_edit(request, pk):
    product = Product.objects.get(id=pk)
    product_serialized = ProductSerializer(product)
    return Response(product_serialized.data)


@api_view(['PUT'])
@authentication_classes([CustomTokenAuthentication])
@permission_classes([IsAuthenticated])
def product_update(request, pk):
    category = Category.objects.get(id=request.data['category_id'])
    product = Product.objects.get(id=pk)
    old_image = product.image.public_id
    if 'image' in request.FILES:
        file = request.FILES['image']
        result = cloudinary.uploader.upload(file, folder='sstore_products')
        if old_image != 'sstore_products/0c32b31941863a0f1fb8e97eaf55f595_lc10im':
            cloudinary.uploader.destroy(old_image)
    else:
        result = {'public_id': old_image}
    request.data._mutable = True
    request.data['image'] = result['public_id']
    request.data._mutable = False
    product_serialized = ProductSerializer(product, data=request.data)
    if product_serialized.is_valid():
        product_serialized.validated_data['category_id'] = category
        product_serialized.save()
        return Response({'message': 'Product updated successfully'})
    return Response(product_serialized.errors, status=400)


@api_view(['DELETE'])
@authentication_classes([CustomTokenAuthentication])
@permission_classes([IsAuthenticated])
def product_delete(request, pk):
    product = Product.objects.get(id=pk)
    product.delete()
    return Response({'message': 'Product deleted successfully'})


@api_view(['DELETE'])
@authentication_classes([CustomTokenAuthentication])
@permission_classes([IsAuthenticated])
def product_delete_several(request):
    ids = request.data['ids']
    Product.objects.filter(id__in=ids).delete()
    return Response({'message': 'Products deleted successfully'})


@api_view(['POST'])
@authentication_classes([CustomTokenAuthentication])
@permission_classes([IsAuthenticated])
def product_create(request):
    try:
        category = Category.objects.get(id=request.data['category_id'])
        if 'image' in request.FILES:
            file = request.FILES['image']
            result = cloudinary.uploader.upload(file, folder='sstore_products')
        else:
            result = {'public_id': 'sstore_products/0c32b31941863a0f1fb8e97eaf55f595_lc10im'}
        product = Product(category_id=category, name=request.data['name'], quantity=request.data['quantity'],
                          quantity_type=request.data['quantity_type'],
                          price_per_quantity=request.data['price_per_quantity'], image=result['public_id'],
                          status=('available' if int(request.data['quantity']) >= 50 else 'few'))
        product.save()
        product_update = ProductUpdate(product_id=product, quantity=product.quantity,
                                       price=request.data['bought_price'], status='added')
        product_update.save()
        return Response({'message': 'Product created and new updated saved successfully'})
    except Exception as e:
        return Response({'error': str(e)}, status=400)


@api_view(['GET'])
@authentication_classes([CustomTokenAuthentication])
@permission_classes([IsAuthenticated])
def products_report(request):
    categories = Category.objects.all().filter(market_id=request.user.id)
    category_serialized = CategorySerializer(categories, many=True)
    markets = [category['id'] for category in category_serialized.data if category['market_id'] == request.user.id]
    products = Product.objects.all().filter(category_id__in=markets)
    products_serialized = ProductSerializer(products, many=True)
    market_name = request.user.market_name
    data = {
        'Mahsulot nomi': [product['name'] for product in products_serialized.data],
        'Kategoriya': [product['category_name'] for product in products_serialized.data],
        'Qoldiq': [product['quantity'] for product in products_serialized.data],
        'Status': [product['status'] for product in products_serialized.data],
        'Narx': [product['price_per_quantity'] for product in products_serialized.data]
    }
    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=3, sheet_name='Products Report')
        worksheet = writer.sheets['Products Report']
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        center_alignment = Alignment(horizontal='center', vertical='center')
        worksheet['A1'] = f"Products Report for {market_name}"
        worksheet['A1'].font = Font(bold=True, size=14, color="000000")
        worksheet['A1'].alignment = Alignment(horizontal='center')
        worksheet.merge_cells('A1:E1')
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        worksheet['A2'] = f"Generated on: {timestamp}"
        worksheet['A2'].font = Font(italic=True, size=10)
        worksheet['A2'].alignment = Alignment(horizontal='center')
        worksheet.merge_cells('A2:E2')
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=4, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = cell_border
        for row in range(5, len(df) + 5):
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = cell_border
                cell.alignment = center_alignment
        for column in range(1, len(data) + 1):
            max_length = max(
                len(str(df.iloc[:, column - 1].values[i]))
                for i in range(len(df)) if df.iloc[:, column - 1].values[i]
            ) + 2
            adjusted_width = min(max_length, 50)
            worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width
    output.seek(0)
    response = HttpResponse(
        content=output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={market_name}_products_report.xlsx'
    return response


@api_view(['POST'])
@authentication_classes([CustomTokenAuthentication])
@permission_classes([IsAuthenticated])
def save_product_updates(request):
    product_updates = request.data
    total_price = 0
    message = 'Product updates saved successfully'
    d = {}
    if product_updates.get('debtor_name'):
        debtor = Debtor.objects.filter(phone=product_updates['debtor_phone'])
        if debtor.exists():
            debtor = debtor.first()
            if debtor.name.lower() != product_updates['debtor_name'].lower():
                return Response({'message': 'Bir raqamdan faqat bitta nomga qarz olish mumkin'})
            message += ' and debt added successfully'
        else:
            market = Market.objects.get(id=request.user.id)
            debtor = Debtor(market_id=market, name=product_updates['debtor_name'],
                            phone=product_updates['debtor_phone'], price=total_price)
            debtor.save()
            message += ' and debtor added successfully'
        d['debtor'] = debtor
    for product_update in product_updates['sells']:
        price = product_update['price'] * product_update['quantity']
        total_price += price
        product = Product.objects.get(id=product_update['product_id'])
        new_product_update = ProductUpdate(product_id=product, quantity=product_update['quantity'], price=price,
                                           debtor=d.get('debtor'))
        new_product_update.save()
        product.quantity -= product_update['quantity']
        if product.quantity < 50:
            product.status = 'few'
        elif product.quantity == 0:
            product.status = 'ended'
        else:
            product.status = 'available'
        product.save()
    if product_updates.get('debtor_name'):
        debtor = Debtor.objects.filter(phone=product_updates['debtor_phone']).first()
        debtor.price += total_price
        debtor.save()
    return Response({'message': message})


@api_view(['POST'])
@authentication_classes([CustomTokenAuthentication])
@permission_classes([IsAuthenticated])
def save_bought_products(request):
    product_updates = request.data
    product = Product.objects.get(id=product_updates['product_id'])
    new_product_update = ProductUpdate(product_id=product, quantity=product_updates['quantity'],
                                       price=product_updates['price'], status='added')
    new_product_update.save()
    product.quantity += product_updates['quantity']
    if product.quantity > 50:
        product.status = 'available'
    elif product.quantity > 0:
        product.status = 'few'
    product.save()
    return Response({'message': 'Product bought successfully'})


@api_view(['GET'])
@authentication_classes([CustomTokenAuthentication])
@permission_classes([IsAuthenticated])
def debtors(request):
    market = Market.objects.get(id=request.user.id)
    debtors = Debtor.objects.filter(market_id=market)
    debtors_serialized = DebtorSerializer(debtors, many=True)
    return Response(debtors_serialized.data)


@api_view(['GET'])
@authentication_classes([CustomTokenAuthentication])
@permission_classes([IsAuthenticated])
def get_debtors_debts(request, pk):
    debtor = Debtor.objects.get(id=pk)
    debtor_serialized = DebtorSerializer(debtor)
    debts_serialized = ProductUpdateSerializer(debtor.debts, many=True)
    return Response({'debtor': debtor_serialized.data, 'debts': debts_serialized.data})


@api_view(['DELETE'])
@authentication_classes([CustomTokenAuthentication])
@permission_classes([IsAuthenticated])
def delete_debt(request, pk):
    debt = ProductUpdate.objects.get(id=pk)
    debtor = debt.debtor
    debt.debtor = None
    debtor.price -= debt.price
    debt.save()
    if float(debtor.price) <= 0:
        debtor.delete()
    else:
        debtor.save()
    return Response({'message': 'Debt deleted successfully'})


@api_view(['GET'])
@authentication_classes([CustomTokenAuthentication])
@permission_classes([IsAuthenticated])
def expenses(request):
    current_month = datetime.now().month
    start_date = timezone.make_aware(datetime(datetime.now().year, current_month, 1))
    expenses = Expense.objects.all().filter(market_id=request.user.id, date__gte=start_date)
    expenses_serialized = ExpenseSerializer(expenses, many=True)
    return Response(expenses_serialized.data)


@api_view(['GET'])
@authentication_classes([CustomTokenAuthentication])
@permission_classes([IsAuthenticated])
def expense_types(request):
    options = dict(Expense.EXPENSE_TYPES)
    return Response(options)


@api_view(['POST'])
@authentication_classes([CustomTokenAuthentication])
@permission_classes([IsAuthenticated])
def expense_save(request):
    market = Market.objects.get(id=request.user.id)
    expense = Expense(market_id=market, type=request.data['type'], price=request.data['price'])
    expense.save()
    return Response({'message': 'Expense saved successfully'})


@api_view(['GET'])
@authentication_classes([CustomTokenAuthentication])
@permission_classes([IsAuthenticated])
def history(request):
    data = request.data
    if data.get('start_month') and data.get('start_day'):
        start_month = data['start_month']
        start_day = data['start_day']
    else:
        start_month = datetime.now().month
        start_day = 1
    start_date = timezone.make_aware(datetime(datetime.now().year, start_month, start_day))
    if data.get('end_month') and data.get('end_day'):
        end_month = data['end_month']
        end_day = data['end_day']
    else:
        end_month = datetime.now().month
        end_day = datetime.now().day
    end_date = timezone.make_aware(datetime(datetime.now().year, end_month, end_day))
    market = Market.objects.get(id=request.user.id)
    categories = Category.objects.all().filter(market_id=market)
    products = Product.objects.all().filter(category_id__in=[category.id for category in categories])
    updates = ProductUpdate.objects.all().filter(product_id__in=[product.id for product in products], status='subed',
                                                 date__gte=start_date, date__lte=end_date)
    updates_serialized = ProductUpdateSerializer(updates, many=True)
    return Response(updates_serialized.data)


@api_view(['DELETE'])
@authentication_classes([CustomTokenAuthentication])
@permission_classes([IsAuthenticated])
def history_delete(request, pk):
    update = ProductUpdate.objects.get(id=pk)
    product = Product.objects.get(id=update.product_id.id)
    product.quantity += update.quantity
    product.save()
    if update.debtor:
        debtor = Debtor.objects.get(id=update.debtor.id)
        debtor.price -= update.price
        debtor.save()
    update.delete()
    return Response({'message': 'History deleted successfully'})


@api_view(['GET'])
@authentication_classes([CustomTokenAuthentication])
@permission_classes([IsAuthenticated])
def history_edit(request, pk):
    update = ProductUpdate.objects.get(id=pk)
    update_serialized = ProductUpdateSerializer(update)
    return Response(update_serialized.data)


@api_view(['PUT'])
@authentication_classes([CustomTokenAuthentication])
@permission_classes([IsAuthenticated])
def history_update(request, pk):
    update = ProductUpdate.objects.get(id=pk)
    product = Product.objects.get(id=update.product_id.id)
    product.quantity += update.quantity
    if update.debtor:
        debtor = Debtor.objects.get(id=update.debtor.id)
        debtor.price -= update.price
        debtor.price += request.data['price']
        debtor.save()
    update.quantity = request.data['quantity']
    update.price = request.data['price']
    product.quantity -= update.quantity
    product.save()
    update.save()
    return Response({'message': 'History updated successfully'})


@api_view(['GET'])
def check(request):
    return Response({'message': 'Server is running'})

# from channels.layers import get_channel_layer
# from asgiref.sync import async_to_sync
#
# def trigger_dashboard_update(user_id):
#     channel_layer = get_channel_layer()
#     async_to_sync(channel_layer.group_send)(
#         f"dashboard_{user_id}",
#         {
#             "type": "dashboard_update"
#         }
#     )
